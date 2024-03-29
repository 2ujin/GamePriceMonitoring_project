# git - https://github.com/2ujin/theteams_project
# language - azure, azure storage, excel, nodejs, python, html, css, javascript
# 2019.07.27 - 2019.08.01
# 30분에 한번씩 가격을 크롤링 (azure - timetrigger 함수 사용)
# 크롤링 한 값 -> azure storage에 저장
# azure storage -> excel로 export
# excel에 저장되어 있는 가격 값을 python 배열에 저장
# python flask를 사용해서 chart html에 값 넘겨서 차트완성 !!!! 

from flask import Flask, render_template, request
from openpyxl import load_workbook
app = Flask(__name__)

load_wb = load_workbook("/Users/Mirim/Desktop/priceTbl.xlsx", data_only=True) #엑셀 파일 갖고오기
load_ws = load_wb['priceTbl'] # 시트 이름으로 불러오기

#30분 주기로 크롤링 한 가격 값
data_arr1 = [load_ws.cell(2, 4).value, load_ws.cell(2, 5).value, load_ws.cell(2, 6).value, load_ws.cell(2, 7).value, load_ws.cell(2, 8).value]; #1
data_arr2 = [load_ws.cell(3, 4).value, load_ws.cell(3, 5).value, load_ws.cell(3, 6).value, load_ws.cell(3, 7).value, load_ws.cell(3, 8).value];
data_arr3 = [load_ws.cell(4, 4).value, load_ws.cell(4, 5).value, load_ws.cell(4, 6).value, load_ws.cell(4, 7).value, load_ws.cell(4, 8).value];
data_arr4 = [load_ws.cell(5, 4).value, load_ws.cell(5, 5).value, load_ws.cell(5, 6).value, load_ws.cell(5, 7).value, load_ws.cell(5, 8).value];
data_arr5 = [load_ws.cell(6, 4).value, load_ws.cell(6, 5).value, load_ws.cell(6, 6).value, load_ws.cell(6, 7).value, load_ws.cell(6, 8).value];
data_arr6 = [load_ws.cell(7, 4).value, load_ws.cell(7, 5).value, load_ws.cell(7, 6).value, load_ws.cell(7, 7).value, load_ws.cell(7, 8).value];
data_arr7 = [load_ws.cell(8, 4).value, load_ws.cell(8, 5).value, load_ws.cell(8, 6).value, load_ws.cell(8, 7).value, load_ws.cell(8, 8).value];
data_arr8 = [load_ws.cell(9, 4).value, load_ws.cell(9, 5).value, load_ws.cell(9, 6).value, load_ws.cell(9, 7).value, load_ws.cell(9, 8).value];
data_arr9 = [load_ws.cell(10, 4).value, load_ws.cell(10, 5).value, load_ws.cell(10, 6).value, load_ws.cell(10, 7).value, load_ws.cell(10, 8).value];
data_arr10 = [load_ws.cell(11, 4).value, load_ws.cell(11, 5).value, load_ws.cell(11, 6).value, load_ws.cell(11, 7).value, load_ws.cell(11, 8).value];

date1 = [load_ws.cell(2, 2).value[20:24], load_ws.cell(3, 2).value[19:24], load_ws.cell(4, 2).value[20:24], load_ws.cell(5, 2).value[19:24], load_ws.cell(6, 2).value[20:24],
         load_ws.cell(7, 2).value[19:24], load_ws.cell(8, 2).value[20:24], load_ws.cell(9, 2).value[19:24], load_ws.cell(10, 2).value[20:24], load_ws.cell(11, 2).value[19:24]];


@app.route('/') #test라는 루트
def main():
    return render_template('game_chart.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5,  data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)

@app.route('/game1.html')
def game1():
    return render_template('game1.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5,  data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)

@app.route('/game2.html')
def game2():
    return render_template('game2.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5, data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)

@app.route('/game3.html')
def game3():
    return render_template('game3.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5, data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)


@app.route('/game4.html')
def game4():
    return render_template('game4.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5, data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)

@app.route('/game5.html')
def game5():
    return render_template('game5.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5, data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)

@app.route('/game_chart.html')
def game_chart():
    return render_template('game_chart.html',
    data_arr1=data_arr1, data_arr2=data_arr2, data_arr3=data_arr3, data_arr4=data_arr4, data_arr5=data_arr5, data_arr6=data_arr6, data_arr7=data_arr7, data_arr8=data_arr8, data_arr9=data_arr9, data_arr10=data_arr10, date1=date1)

if __name__ == '__main__':
    app.run()