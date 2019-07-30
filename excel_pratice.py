# git - https://github.com/2ujin/theteams_project
# language - azure, azure storage, excel, nodejs, python, html, css, javascript
# 2019.07.27 - 2019.08.01
# 30분에 한번씩 가격을 크롤링 (azure - timetrigger 함수 사용)
# 크롤링 한 값 -> azure storage에 저장
# azure storage -> excel로 export
# python을 사용해서 배열에 가격 정보를 저장
# python -> html로 리턴해서 chartjs에 가격 정보 insert

from openpyxl import load_workbook

load_wb = load_workbook("/Users/s2017/Desktop/priceTbl.xlsx", data_only=True) #엑셀 파일 갖고오기
load_ws = load_wb['priceTbl.'] # 시트 이름으로 불러오기

#30분 주기로 크롤링 한 가격 값
data_arr1 = [load_ws.cell(2, 4).value, load_ws.cell(2, 5).value, load_ws.cell(2, 6).value];
data_arr2 = [load_ws.cell(3, 4).value, load_ws.cell(3, 5).value, load_ws.cell(3, 6).value];
data_arr3 = [load_ws.cell(4, 4).value, load_ws.cell(4, 5).value, load_ws.cell(4, 6).value];
data_arr4 = [load_ws.cell(5, 4).value, load_ws.cell(5, 5).value, load_ws.cell(5, 6).value];
data_arr5 = [load_ws.cell(6, 4).value, load_ws.cell(6, 5).value, load_ws.cell(6, 6).value];
data_arr6 = [load_ws.cell(7, 4).value, load_ws.cell(7, 5).value, load_ws.cell(7, 6).value];
data_arr7 = [load_ws.cell(8, 4).value, load_ws.cell(8, 5).value, load_ws.cell(8, 6).value];
data_arr8 = [load_ws.cell(9, 4).value, load_ws.cell(9, 5).value, load_ws.cell(9, 6).value];
data_arr9 = [load_ws.cell(10, 4).value, load_ws.cell(10, 5).value, load_ws.cell(10, 6).value];
data_arr10 = [load_ws.cell(11, 4).value, load_ws.cell(11, 5).value, load_ws.cell(11, 6).value];

print(data_arr10[0]);
print(data_arr10[1]);
print(data_arr10[2]);


